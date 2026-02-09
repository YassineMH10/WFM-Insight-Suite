import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from datetime import datetime, date, time

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Planning & Assiduité", layout="wide")

# ----------------------------
# Footer badge
# ----------------------------
def footer():
    st.markdown(
        """
        <div style="margin-top:18px; padding-top:10px; border-top:1px solid #E6E6E6;
                    color:#6c757d; font-size:12px; text-align:center;">
            Developed by <b>MAHAMID Yassine</b>
        </div>
        """,
        unsafe_allow_html=True
    )

# ----------------------------
# Helpers (PDF)
# ----------------------------
TIME_PATTERN = r"\d{2}:\d{2}"

def get_duration(a, b):
    try:
        t1 = datetime.strptime(a, "%H:%M")
        t2 = datetime.strptime(b, "%H:%M")
        return (t2 - t1).seconds // 60
    except:
        return 0

def is_continuation_line(line, times):
    return len(times) == 2 and len(line.split()) <= 3

def extract_planning_from_pdf(uploaded_pdf) -> pd.DataFrame:
    rows = []
    last_index = -1

    with pdfplumber.open(BytesIO(uploaded_pdf.getvalue())) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for line in text.split("\n"):
                line = line.strip()
                times = re.findall(TIME_PATTERN, line)

                if line.lower().startswith(("cré", "le ")):
                    continue

                if last_index >= 0 and is_continuation_line(line, times):
                    a, b = times
                    dur = get_duration(a, b)
                    if dur <= 15:
                        if rows[last_index][5] == "":
                            rows[last_index][5] = f"{a} - {b}"
                    elif dur >= 30:
                        rows[last_index][4] = f"{a} - {b}"
                    continue

                if len(times) < 2:
                    continue

                parts = line.split()
                matricule = parts[0]
                if str(matricule).lower().startswith("cré"):
                    continue

                first_time = times[0]
                name = line.split(matricule)[1].split(first_time)[0].strip()

                pause1 = ""
                pause2 = ""
                repas = ""

                blocks = [(times[i], times[i+1]) for i in range(1, len(times)-1, 2)]
                pauses = []

                for a, b in blocks:
                    dur = get_duration(a, b)
                    if dur <= 15:
                        pauses.append(f"{a} - {b}")
                    elif dur >= 30:
                        repas = f"{a} - {b}"

                if len(pauses) >= 1:
                    pause1 = pauses[0]
                if len(pauses) >= 2:
                    pause2 = pauses[1]

                rows.append([
                    matricule,
                    name,
                    times[0],
                    pause1,
                    repas,
                    pause2,
                    times[-1]
                ])
                last_index += 1

    # WFM validation (break/lunch)
    for r in rows:
        h_debut = r[2]

        if r[3] and get_duration(*r[3].split(" - ")) > 15:
            r[4] = r[3]
            r[3] = ""

        if r[5] and get_duration(*r[5].split(" - ")) > 15:
            r[4] = r[5] if not r[4] else r[4] + " / " + r[5]
            r[5] = ""

        if r[4]:
            repas_start = r[4].split(" - ")[0]
            if get_duration(h_debut, repas_start) < 60:
                r[4] = ""

    df_planning_pdf = pd.DataFrame(rows, columns=[
        "Matricule","Nom","Heure de début",
        "Pause courte 1","Repas","Pause courte 2","Heure de fin"
    ])

    # WFM READY
    df_planning_pdf = df_planning_pdf[df_planning_pdf["Nom"].astype(str).str.match(r"^[A-Za-zÀ-ÖØ-öø-ÿ]")]
    df_planning_pdf.drop(columns=["Nom"], inplace=True)

    df_planning_pdf.rename(columns={
        "Heure de début": "Heure de départ",
        "Heure de fin": "Horaire de fin",
        "Pause courte 1": "Break 1_D",
        "Pause courte 2": "Break 2_D",
        "Repas": "Lunch D"
    }, inplace=True)

    def insert_after(df, col, new):
        i = df.columns.get_loc(col) + 1
        df.insert(i, new, "")

    insert_after(df_planning_pdf, "Break 1_D", "Break 1_F")
    insert_after(df_planning_pdf, "Lunch D", "Lunch F")
    insert_after(df_planning_pdf, "Break 2_D", "Break 2_F")

    def split_range(v):
        if pd.isna(v) or v == "":
            return "", ""
        p = re.split(r"\s*-\s*", str(v))
        return (p[0], p[1]) if len(p) == 2 else (v, "")

    df_planning_pdf["Break 1_D"], df_planning_pdf["Break 1_F"] = zip(*df_planning_pdf["Break 1_D"].map(split_range))
    df_planning_pdf["Lunch D"], df_planning_pdf["Lunch F"] = zip(*df_planning_pdf["Lunch D"].map(split_range))
    df_planning_pdf["Break 2_D"], df_planning_pdf["Break 2_F"] = zip(*df_planning_pdf["Break 2_D"].map(split_range))

    def to_min(t):
        if t == "" or pd.isna(t): return None
        h, m = map(int, t.split(":"))
        return h*60 + m

    hours = []
    for _, r in df_planning_pdf.iterrows():
        s, e = to_min(r["Heure de départ"]), to_min(r["Horaire de fin"])
        ld, lf = to_min(r["Lunch D"]), to_min(r["Lunch F"])
        if s is None or e is None:
            hours.append("")
            continue
        if e <= s: e += 1440
        lunch = 0
        if ld and lf:
            if lf <= ld: lf += 1440
            lunch = lf - ld
        hours.append(round((e - s - lunch) / 60, 2))

    df_planning_pdf.insert(df_planning_pdf.columns.get_loc("Horaire de fin") + 1, "Heures planifiées", hours)
    return df_planning_pdf


# ----------------------------
# Helpers COMPO
# ----------------------------
def read_compo(uploaded_excel) -> pd.DataFrame:
    xl = pd.ExcelFile(BytesIO(uploaded_excel.getvalue()))
    for sheet in xl.sheet_names:
        df_sheet = xl.parse(sheet)
        cols = [str(c).strip().lower() for c in df_sheet.columns]
        if any("matricule rh" in c for c in cols) and any("log téléphonie1" in c or "log telephonie1" in c for c in cols):
            return df_sheet.copy()
    raise ValueError("COMPO/KPI file not recognized. Verify headers (Matricule RH, Log Téléphonie1, Nom Agent, File, Tls, OPS).")


# ----------------------------
# Helpers HERMES (.xls)
# ----------------------------
def convert_time_format(time_str):
    try:
        return datetime.strptime(time_str, "%Hh%M").strftime("%I:%M:%S %p")
    except ValueError:
        try:
            return datetime.strptime(time_str, "%I:%M %p").strftime("%I:%M:%S %p")
        except ValueError:
            return time_str

def extract_agent_data_from_hermes_xls(uploaded_xls, report_date_str: str) -> pd.DataFrame:
    # read xls (skiprows=5) then convert to csv-like dataframe in memory
    xls_data = pd.read_excel(BytesIO(uploaded_xls.getvalue()), skiprows=5, engine="xlrd")
    data = xls_data.copy()

    arrival_rows = data[data.apply(lambda row: row.astype(str).str.contains("Arrivée-Départ :", na=False).any(), axis=1)]

    results = []
    for idx, row in arrival_rows.iterrows():
        first_val = row.dropna().values[0]
        time_info = str(first_val).split("Arrivée-Départ :")[-1].strip()
        if "-" not in time_info:
            continue

        connection_time, disconnection_time = time_info.split("-", 1)
        connection_time = convert_time_format(connection_time.strip())
        disconnection_time = convert_time_format(disconnection_time.strip())

        for preceding_idx in range(idx - 10, -1, -1):
            agent_info = data.iloc[preceding_idx, 0]
            if pd.notna(agent_info):
                s = str(agent_info)
                if ":" in s and "," in s:
                    matricule, _name = s.split(":", 1)
                    try:
                        log = int(str(matricule).strip())
                    except:
                        continue
                    results.append({
                        "DATE": report_date_str,
                        "Log Hermes": log,
                        "Arrivé": connection_time,
                        "Départ": disconnection_time
                    })
                    break

    return pd.DataFrame(results)


# ----------------------------
# Excel writer (styled)
# ----------------------------
def write_sheet(wb, name, df_in):
    ws = wb.create_sheet(name)

    header_fill = PatternFill("solid", fgColor="002060")
    red_fill = PatternFill("solid", fgColor="C00000")
    beige_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    redlight_fill = PatternFill("solid", fgColor="F4CCCC")
    blue_fill = PatternFill("solid", fgColor="002060")

    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    font_white = Font(color="FFFFFF", bold=True)
    font_black = Font(color="000000")
    font_green = Font(color="38761D")
    font_red = Font(color="CC0000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(dataframe_to_rows(df_in, index=False, header=True), start=1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(i, j, val)
            cell.alignment = center
            cell.border = border

            if i == 1:
                cell.fill = header_fill
                cell.font = font_white
            elif val == "Total":
                cell.fill = blue_fill
                cell.font = font_white
            elif j == 1:
                cell.fill = red_fill
                cell.font = font_white
            elif df_in.columns[j-1] in ["Heures planifiées", "Ecart Retards", "Heure absence"]:
                cell.fill = beige_fill
                cell.font = font_black
            elif df_in.columns[j-1] in ["%Absence", "%Retards", "%Assiduite"]:
                try:
                    pct = float(str(val).replace('%', '').replace(',', '.'))
                    cell.font = font_green if pct <= 5 else font_red
                    cell.fill = green_fill if pct <= 5 else redlight_fill
                except:
                    pass

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2


# ----------------------------
# Email HTML blocks
# ----------------------------
def df_to_html_table(df_in, title):
    cols = df_in.columns.tolist()
    header_bg = "#203864"
    header_fg = "#FFFFFF"
    row_alt1 = "#F7F8FA"
    row_alt2 = "#FFFFFF"

    html = f"""
    <div style="margin-top:14px; margin-bottom:10px;">
      <div style="font-weight:bold; font-size:14px; color:#203864; margin-bottom:4px;">
        {title}
      </div>
      <table cellpadding="4" cellspacing="0"
             style="border-collapse:collapse; width:100%; max-width:100%;
                    font-family:Calibri, Arial, sans-serif;
                    font-size:13px; border:1px solid #D0D0D0;">
        <thead>
          <tr style="background:{header_bg}; color:{header_fg}; text-align:center;">
    """
    for c in cols:
        html += f'<th style="border:1px solid #D0D0D0; padding:6px 4px; word-break:break-word;">{c}</th>'
    html += "</tr></thead><tbody>"

    for i, (_, row) in enumerate(df_in.iterrows()):
        bg = row_alt1 if i % 2 == 0 else row_alt2
        html += f'<tr style="background:{bg}; color:#000000; text-align:center;">'
        for c in cols:
            val = row[c]
            display_val = "" if pd.isna(val) else str(val)
            cell_style = "border:1px solid #D0D0D0; padding:6px 4px; word-break:break-word;"

            if (c in ["Equipe", "File"] and str(val) == "Total"):
                cell_style += "background:#D9E1F2; font-weight:bold;"

            if c in ["%Absence", "%Retards", "%Assiduite"] and isinstance(display_val, str) and "%" in display_val:
                try:
                    num = float(display_val.replace("%", "").replace(",", "."))
                    if c == "%Assiduite":
                        if num <= 5:
                            cell_style += "background:#D4EFDF; color:#145A32; font-weight:bold;"
                        elif num <= 10:
                            cell_style += "background:#FCF3CF; color:#7D6608; font-weight:bold;"
                        else:
                            cell_style += "background:#F5B7B1; color:#922B21; font-weight:bold;"
                except:
                    pass

            html += f'<td style="{cell_style}">{display_val}</td>'
        html += "</tr>"

    html += "</tbody></table></div>"
    return html

def copy_buttons(subject: str, html_body: str):
    component = f"""
    <div style="font-family:Calibri, Arial; margin-top:10px;">
      <div style="margin-bottom:6px; font-weight:bold;">📋 Copier vers Gmail :</div>

      <textarea id="sbj" rows="2" style="width:100%;">{subject}</textarea>
      <button style="margin-top:6px; padding:8px 14px; border:0; border-radius:6px; cursor:pointer; background:#28A745; color:white;"
              onclick="navigator.clipboard.writeText(document.getElementById('sbj').value)">
        ✅ Copier l'objet
      </button>

      <div style="height:10px;"></div>

      <div id="htmlBody" style="display:none;">{html_body}</div>
      <button style="padding:8px 14px; border:0; border-radius:6px; cursor:pointer; background:#0078D7; color:white;"
              onclick="navigator.clipboard.write([
                new ClipboardItem({{
                  'text/html': new Blob([document.getElementById('htmlBody').innerHTML], {{type:'text/html'}})
                }})
              ])">
        ✅ Copier le corps HTML (coller formaté dans Gmail)
      </button>
    </div>
    """
    st.components.v1.html(component, height=210, scrolling=False)


# ----------------------------
# Assiduité core
# ----------------------------
def parse_time_user(s: str):
    s = str(s).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            pass
    raise ValueError("Invalid time format. Use HH:MM or HH:MM:SS (e.g., 11:00).")

def hhmm_to_minutes(val):
    if pd.isna(val):
        return 0
    if isinstance(val, (int, float)):
        return int(float(val) * 60)
    try:
        h, m = map(int, str(val).split(':')[:2])
        return h * 60 + m
    except Exception:
        return 0

def format_duree(minutes):
    if pd.isna(minutes): return "00:00:00"
    h, m = divmod(int(minutes), 60)
    return f"{h}:{m:02}:00"

def prepare_resume(df_in, group_col):
    grp = df_in.groupby(group_col, dropna=False).agg({
        "Shift Duration": "sum",
        "Retard (min)": "sum",
        "Absent (min)": "sum"
    }).reset_index()

    grp["%Absence"] = grp["Absent (min)"] / grp["Shift Duration"] * 100
    grp["%Retards"] = grp["Retard (min)"] / grp["Shift Duration"] * 100
    grp["%Assiduite"] = grp["%Absence"] + grp["%Retards"]

    grp["Heures planifiées"] = grp["Shift Duration"].apply(format_duree)
    grp["Ecart Retards"] = grp["Retard (min)"].apply(format_duree)
    grp["Heure absence"] = grp["Absent (min)"].apply(format_duree)

    grp["%Absence"] = grp["%Absence"].fillna(0).map(lambda x: f"{x:.1f}%")
    grp["%Retards"] = grp["%Retards"].fillna(0).map(lambda x: f"{x:.1f}%")
    grp["%Assiduite"] = grp["%Assiduite"].fillna(0).map(lambda x: f"{x:.1f}%")

    resume = grp[[group_col, "Heures planifiées", "Ecart Retards", "Heure absence",
                  "%Absence", "%Retards", "%Assiduite"]]
    resume.columns = ["Equipe" if group_col == "Tls" else "File"] + list(resume.columns[1:])

    total_plan = grp["Shift Duration"].sum()
    total_ret = grp["Retard (min)"].sum()
    total_abs = grp["Absent (min)"].sum()
    total_ass = (total_ret + total_abs) / total_plan * 100 if total_plan else 0

    resume.loc[len(resume)] = [
        "Total",
        format_duree(total_plan),
        format_duree(total_ret),
        format_duree(total_abs),
        f"{(total_abs / total_plan * 100):.1f}%" if total_plan else "0.0%",
        f"{(total_ret / total_plan * 100):.1f}%" if total_plan else "0.0%",
        f"{total_ass:.1f}%"
    ]
    return resume


# ============================================================
# UI
# ============================================================
st.title("📄 Planning / Assiduité – Générateur Excel + Email (HTML)")

with st.sidebar:
    st.header("Configuration")
    mode = st.radio(
        "Choose output",
        ["Planning with Compo only", "Assiduity report only", "Generate both outputs"],
        index=2
    )

    signature_name = st.text_input("Signature (email)", value="Yassine MAHAMID")
    project_name = st.text_input("Project name", value="CNSS")

    report_date = st.text_input("Report date (dd/mm)", value=datetime.today().strftime("%d/%m"))
    time_cutoff = st.text_input("Status time (e.g., 10:30)", value="10:30")

    apply_filter = st.checkbox("Apply start-time filter (Heure de départ <= limit)?", value=False)
    limit_str = st.text_input("Time limit (HH:MM)", value="11:00") if apply_filter else None

    st.markdown("---")
    st.subheader("Files upload")
    pdf_file = st.file_uploader("Teleopti Planning PDF", type=["pdf"])
    compo_file = st.file_uploader("COMPO/KPI Excel", type=["xlsx", "xls"])
    hermes_file = st.file_uploader("HERMES raw export (.xls)", type=["xls"])

need_planning = mode in ["Planning with Compo only", "Generate both outputs"]
need_assiduite = mode in ["Assiduity report only", "Generate both outputs"]

# Basic checks
if not pdf_file or not compo_file:
    st.info("Upload at least: PDF + COMPO/KPI to start.")
    footer()
    st.stop()

if need_assiduite and not hermes_file:
    st.info("Assiduity mode requires: PDF + COMPO/KPI + HERMES (.xls).")
    footer()
    st.stop()

# Process button
run = st.button("🚀 Generate outputs", type="primary")

if not run:
    footer()
    st.stop()

# ============================================================
# Processing
# ============================================================
try:
    df_planning = extract_planning_from_pdf(pdf_file)
    compo_df = read_compo(compo_file)

    planning_df = df_planning.copy()
    planning_df["Matricule"] = planning_df["Matricule"].astype(str).str.strip()
    compo_df["Matricule RH"] = compo_df["Matricule RH"].astype(str).str.strip()

except Exception as e:
    st.error(f"Error while processing PDF/COMPO: {e}")
    footer()
    st.stop()

# Output 1: Planning + Compo
if need_planning:
    st.subheader("✅ Planning with Compo")

    planning_cols = planning_df.columns.difference(["Matricule"], sort=False)
    merged_planning_df = pd.merge(
        compo_df,
        planning_df[["Matricule"] + list(planning_cols)],
        left_on="Matricule RH",
        right_on="Matricule",
        how="left"
    ).drop(columns="Matricule")

    st.dataframe(merged_planning_df, use_container_width=True, height=380)

    # Build Excel in memory
    wb_pl = Workbook()
    ws_pl = wb_pl.active
    ws_pl.title = "Compo with planning"

    for row in dataframe_to_rows(merged_planning_df, index=False, header=True):
        ws_pl.append(row)

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    blue_fill = PatternFill(start_color="000066", end_color="000066", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    black_font = Font(color="000000")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    header = [cell.value for cell in ws_pl[1]]
    kpi_cols = ["Matricule RH", "Log Téléphonie1", "Nom Agent", "File", "Tls", "OPS"]
    planning_cols_style = list(merged_planning_df.columns.difference(kpi_cols, sort=False))

    for col_idx, col_name in enumerate(header, 1):
        col_letter = get_column_letter(col_idx)
        for row_idx, cell in enumerate(ws_pl[col_letter], 1):
            cell.alignment = center_alignment
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = red_fill
                cell.font = white_font
            elif col_name in planning_cols_style:
                cell.fill = white_fill
                cell.font = black_font
            elif col_name in kpi_cols:
                cell.fill = blue_fill
                cell.font = white_font

    for col in ws_pl.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_pl.column_dimensions[col[0].column_letter].width = max_len + 4

    out_buf = BytesIO()
    wb_pl.save(out_buf)
    out_buf.seek(0)

    # filename
    safe_date = report_date.replace("/", "-")
    file_name_planning = f"Planning_{project_name}_{safe_date}.xlsx"

    st.download_button(
        "⬇️ Download Planning + Compo Excel",
        data=out_buf.getvalue(),
        file_name=file_name_planning,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Output 2: Assiduité
if need_assiduite:
    st.subheader("✅ Assiduity report")

    try:
        # Hermes extraction
        report_date_iso = datetime.today().strftime("%Y-%m-%d")
        hermes_df = extract_agent_data_from_hermes_xls(hermes_file, report_date_iso)
        if hermes_df.empty:
            raise ValueError("No records extracted from HERMES. Verify .xls includes 'Arrivée-Départ :'.")

        # Merge base
        df_ass = pd.merge(
            compo_df,
            planning_df[["Matricule", "Heure de départ", "Horaire de fin", "Heures planifiées"]],
            left_on="Matricule RH",
            right_on="Matricule",
            how="left"
        ).drop(columns="Matricule")

        hermes_df["Log Hermes"] = hermes_df["Log Hermes"].astype(str).str.strip()
        df_ass["Log Téléphonie1"] = df_ass["Log Téléphonie1"].astype(str).str.strip()

        df = pd.merge(
            df_ass,
            hermes_df[["Log Hermes", "Arrivé"]],
            left_on="Log Téléphonie1",
            right_on="Log Hermes",
            how="left"
        ).drop(columns="Log Hermes")

        df = df[~df["File"].astype(str).str.contains("ALD", na=False)]

        df["Heure de départ"] = pd.to_datetime(df["Heure de départ"], format="%H:%M", errors="coerce")
        df["Horaire de fin"] = pd.to_datetime(df["Horaire de fin"], format="%H:%M", errors="coerce")
        df["Arrivé"] = pd.to_datetime(df["Arrivé"], format="%I:%M:%S %p", errors="coerce")

        if apply_filter and limit_str:
            heure_limite = parse_time_user(limit_str)
            df["_HeureDepart_time"] = df["Heure de départ"].dt.time
            df = df[df["_HeureDepart_time"].notna() & (df["_HeureDepart_time"] <= heure_limite)].copy()
            df.drop(columns=["_HeureDepart_time"], inplace=True, errors="ignore")

        df["Shift Duration"] = df["Heures planifiées"].apply(hhmm_to_minutes)

        df["Retard (min)"] = np.where(
            df["Arrivé"] > df["Heure de départ"],
            (df["Arrivé"] - df["Heure de départ"]).dt.total_seconds() / 60,
            0
        )
        df["Absent (min)"] = np.where(
            df["Arrivé"].isna() & df["Heure de départ"].notna(),
            df["Shift Duration"],
            0
        )

        rapport_tls = prepare_resume(df, "Tls")
        rapport_file = prepare_resume(df, "File")

        df_retard = df[df["Retard (min)"] > 0].copy()
        df_retard["Arrivé"] = df_retard["Arrivé"].dt.strftime("%H:%M")
        df_retard["Heure de départ"] = df_retard["Heure de départ"].dt.strftime("%H:%M")
        df_retard["Retard"] = df_retard["Retard (min)"].apply(format_duree)

        df_absents = df[df["Absent (min)"] > 0].copy()
        df_absents["Arrivé"] = "ABS"
        df_absents["Heure de départ"] = df_absents["Heure de départ"].dt.strftime("%H:%M")

        # Build Excel output
        wb = Workbook()
        ws0 = wb.active
        wb.remove(ws0)

        write_sheet(wb, "Rapport Tls", rapport_tls)
        write_sheet(wb, "Agents en Retard", df_retard)
        write_sheet(wb, "Agents Absents", df_absents)
        write_sheet(wb, "Rapport par File", rapport_file)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        date_str_for_file = report_date.replace("/", "-")
        final_name = f"Etat_d_assiduite_{date_str_for_file}_{project_name}.xlsx"

        st.download_button(
            "⬇️ Download Assiduity Excel",
            data=buf.getvalue(),
            file_name=final_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Email
        total_plan = df["Shift Duration"].sum()
        total_ret = df["Retard (min)"].sum()
        total_abs = df["Absent (min)"].sum()

        TA_abs = (total_abs / total_plan * 100) if total_plan else 0
        TA_ret = (total_ret / total_plan * 100) if total_plan else 0
        TA_global = TA_abs + TA_ret

        TA_abs_txt = f"{TA_abs:.1f}".replace(".", ",")
        TA_ret_txt = f"{TA_ret:.1f}".replace(".", ",")
        TA_global_txt = f"{TA_global:.1f}".replace(".", ",")

        nb_planifies = df[df["Shift Duration"] > 0]["Log Téléphonie1"].nunique()
        nb_absents = df_absents["Log Téléphonie1"].nunique()

        html_table_file = df_to_html_table(rapport_file, "Assiduity breakdown by skill (File)")
        html_table_tls = df_to_html_table(rapport_tls, "Assiduity breakdown by team (TL)")

        email_subject = f"{project_name} ==> Etat d'assiduité du {report_date} arrêté à {time_cutoff} (TA {TA_global_txt} %)"

        email_html = f"""
        <div style="background:#F4F6F9; padding:12px 0;">
          <div style="max-width:800px; margin:0 auto; background:#FFFFFF;
                      padding:16px 16px 18px 16px; border-radius:6px;
                      border:1px solid #D0D7E2;
                      font-family:Calibri, Arial, sans-serif;
                      font-size:14px; color:#1F1F1F; line-height:1.5;">

            <div style="border-left:4px solid #1F4E79; padding-left:10px; margin-bottom:12px;">
              <div style="font-size:16px; font-weight:bold; color:#1F4E79;">
                Etat d'assiduité – {project_name} – {report_date}
              </div>
              <div style="font-size:12px; color:#6C757D; margin-top:2px;">
                Situation arrêtée à {time_cutoff}
              </div>
            </div>

            <div style="margin-bottom:10px;">
              <span style="display:inline-block; margin-right:6px;
                           background:#FDEDEC; color:#C0392B; padding:4px 8px;
                           border-radius:12px; font-size:12px; font-weight:bold;">
                TA global : {TA_global_txt} %
              </span>
              <span style="display:inline-block; margin-right:6px;
                           background:#FEF5E7; color:#B9770E; padding:4px 8px;
                           border-radius:12px; font-size:12px;">
                Absences : {TA_abs_txt} %
              </span>
              <span style="display:inline-block;
                           background:#E8F8F5; color:#117864; padding:4px 8px;
                           border-radius:12px; font-size:12px;">
                Retards : {TA_ret_txt} %
              </span>
            </div>

            <p style="margin:0 0 8px 0;"><b>Bonjour,</b></p>

            <p style="margin:0 0 10px 0;">
              À <b>{time_cutoff}</b>, sur <b>{nb_planifies} agents planifiés</b>,
              <b style="color:#C00000;">{nb_absents} agents absents</b>.
              Le taux de non-respect de l'assiduité (retards + absences) s'élève à
              <b style="color:#C00000;">{TA_global_txt} %</b>,
              dont <b>{TA_abs_txt} %</b> liés aux absences et <b>{TA_ret_txt} %</b> aux retards.
            </p>

            {html_table_file}
            {html_table_tls}

            <p style="margin:10px 0 14px 0;">
              Le fichier Excel joint reprend le détail des présences, retards et absences
              au niveau agent (log téléphonie et matricule), afin de faciliter le plan d'action
              avec les équipes concernées.
            </p>

            <p style="margin:0;">
              <b>Cordialement,</b><br>
              <b>{signature_name}</b><br>
              Analyste IDP<br>
              Direction WorkForce Management &amp; Reporting<br>
              <span style="color:#6C757D; font-size:12px;">Developed by MAHAMID Yassine</span>
            </p>

          </div>
        </div>
        """

        st.subheader("✉️ Email (HTML)")
        st.text_input("Subject", value=email_subject)
        st.components.v1.html(email_html, height=520, scrolling=True)
        copy_buttons(email_subject, email_html)

    except Exception as e:
        st.error(f"Error in Assiduity process: {e}")

footer()
